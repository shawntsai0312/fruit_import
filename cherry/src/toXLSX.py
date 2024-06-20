import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

def transformDate(date):
    if (date[0] == '0'):
        month = date[1]
    else:
        month = date[:2]
    if (date[2] == '0'):
        day = date[3]
    else:
        day = date[2:]
    return month + '月' + day + '日'


def merge_cells_in_row(sheet, row):
    # 初始化變量
    start_col = None
    previous_value = None

    for col in range(1, sheet.max_column + 1):
        current_value = sheet.cell(row=row, column=col).value
        
        if current_value == previous_value:
            # 如果當前值與前一個值相同，繼續
            continue
        else:
            # 如果當前值與前一個值不同，合併前一組儲存格
            if start_col is not None and start_col < col - 1:
                sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=col-1)
                
            # 設置新的開始列和前一個值
            start_col = col
            previous_value = current_value

    # 最後一組儲存格也需要合併
    if start_col is not None and start_col < sheet.max_column:
        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=sheet.max_column)


def toXLSX(cherrys):
    # Create a directory under categorizedXLSX/
    if not os.path.exists('categorizedXLSX'):
        os.mkdir('categorizedXLSX')

    # 讀取 Excel 模板
    workbook = openpyxl.Workbook()

    for cherry in cherrys:
        # print(cherry)
        if cherry.arrival_date not in workbook.sheetnames:
            # print(cherry.arrival_date, 'not in workbook.sheetnames')
            workbook.create_sheet(cherry.arrival_date, 0)
            sheet = workbook[cherry.arrival_date]
            sheet['A1'] = '台灣萊翁司國際貿易股份有限公司'
            sheet['A3'] = '日期：'
            sheet['B3'] = transformDate(cherry.arrival_date)
            sheet['A4'] = '櫃號：'
            sheet['E3'] = '品名：'
            sheet['E4'] = '專案：'
            sheet['I3'] = 'T：0939-235-899'
            sheet['I4'] = 'F：2339-5044'
            sheet['A6'] = '市場'
            sheet['B6'] = '客戶名稱'
        else:
            sheet = workbook[cherry.arrival_date]

        for cell in sheet[6]:
            if cell.value is None:
                column_letter = get_column_letter(cell.column)
                sheet[f'{column_letter}5'] = cherry.label
                sheet[f'{column_letter}6'] = cherry.variety_code
                sheet[f'{column_letter}7'] = str(cherry.pack_code) + 'K'
                sheet[f'{column_letter}8'] = str(cherry.size_code) + 'R'
                sheet[f'{column_letter}9'] = cherry.ctns
                break
    workbook.remove(workbook['Sheet'])
    for sheet in workbook:
        last_column = 0
        last_column_letter = ''
        for cell in sheet[6]:
            if cell.value is None:
                last_column = cell.column
                # print(last_column)
                if last_column <= 11:  # 11 is the column number of 'K'
                    last_column_letter = get_column_letter(11)
                else:
                    last_column_letter = get_column_letter(last_column)
                sheet[f'{last_column_letter}6'] = '共'
                sheet[f'{last_column_letter}9'] = f'=SUM(C9:{get_column_letter(last_column-1)}9)'
                break
        sheet.merge_cells(f'A1:{last_column_letter}1')
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='標楷體', size=15)
        sheet['A1'].font = Font(name='標楷體', size=30, underline='single')
        sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        merge_cells_in_row(sheet, 5)
        merge_cells_in_row(sheet, 6)
        for sheet in workbook:
            for row in sheet.iter_rows(min_row=5):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')

    workbook.save('categorizedXLSX/output.xlsx')

    # sheet['A1'] = '台灣萊翁司國際貿易股份有限公司'

    # # 假設你要將 CSV 中的數據寫入模板的 A1 開始的區域
    # start_row = 1
    # start_col = 1

    # # # 將 CSV 資料寫入模板
    # # for i, row in df.iterrows():
    # #     for j, value in enumerate(row):
    # #         sheet.cell(row=start_row + i, column=start_col + j, value=value)

    # # 儲存新的 Excel 文件
    # output_file = 'output.xlsx'
    # workbook.save(output_file)

    # print(f"Data has been written to {output_file}")
